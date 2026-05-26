import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
import os

OUTPUT = os.path.join(os.path.dirname(__file__), '..', 'output')
EXCEL_PATH = os.path.join(OUTPUT, 'uk_he_outcomes_analysis.xlsx')

GREEN_FILL  = PatternFill('solid', fgColor='C6EFCE')
AMBER_FILL  = PatternFill('solid', fgColor='FFEB9C')
RED_FILL    = PatternFill('solid', fgColor='FFC7CE')
HEADER_FILL = PatternFill('solid', fgColor='1A3A5C')
HEADER_FONT = Font(color='FFFFFF', bold=True)
BOLD        = Font(bold=True)
CENTRE      = Alignment(horizontal='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTRE
        cell.border = thin_border


def auto_width(ws, min_width=10):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 2, min_width)


def rag_fill(value, thresholds):
    if value is None:
        return None
    if value <= thresholds[0]:
        return GREEN_FILL
    elif value <= thresholds[1]:
        return AMBER_FILL
    else:
        return RED_FILL


wb = openpyxl.Workbook()
wb.remove(wb.active)


# ============================================================
# Sheet 1: Access Summary
# ============================================================
ws1 = wb.create_sheet('Access Summary')

dep = pd.read_csv(os.path.join(OUTPUT, 'access_by_deprivation.csv'))
eth = pd.read_csv(os.path.join(OUTPUT, 'access_by_ethnicity.csv'))

ws1['A1'] = 'ACCESS SUMMARY: HE Progression Rates'
ws1['A1'].font = Font(bold=True, size=13)
ws1.merge_cells('A1:F1')

ws1['A3'] = 'Section 1: Deprivation Gap (FSM Status)'
ws1['A3'].font = BOLD

fsm_row = dep[dep['analysis'] == 'FSM gap (most recent year)'].iloc[0]
headers = ['FSM Status', 'Rates', 'Gap', 'Trend']
for i, h in enumerate(headers, 1):
    ws1.cell(5, i).value = h
style_header_row(ws1, 5, len(headers))
ws1.cell(6, 1).value = 'Non-FSM vs FSM'
ws1.cell(6, 2).value = fsm_row['rates']
ws1.cell(6, 3).value = fsm_row['gap_summary']
ws1.cell(6, 4).value = fsm_row['trend']

ws1['A8'] = 'Section 2: HE Access Rate by IMD Quintile'
ws1['A8'].font = BOLD

imd_rows = dep[dep['analysis'].str.startswith('IMD access rate')].copy()
imd_rows['quintile'] = imd_rows['analysis'].str.replace('IMD access rate: ', '', regex=False)
imd_rows['rate'] = imd_rows['rates'].str.replace('%', '').astype(float)
quintile_headers = ['Quintile', 'Access Rate (%)', 'vs Q1 (pp)']
for i, h in enumerate(quintile_headers, 1):
    ws1.cell(9, i).value = h
style_header_row(ws1, 9, len(quintile_headers))
for r_idx, (_, row) in enumerate(imd_rows.iterrows(), 10):
    ws1.cell(r_idx, 1).value = row['quintile']
    ws1.cell(r_idx, 2).value = row['rate']
    ws1.cell(r_idx, 3).value = row['gap_summary']
    fill = GREEN_FILL if 'Q1' in row['quintile'] else None
    if fill:
        for c in range(1, 4):
            ws1.cell(r_idx, c).fill = fill

ws1['A16'] = 'Section 3: HE Progression by Ethnic Group (top 15)'
ws1['A16'].font = BOLD

eth_headers = ['Ethnic Group', 'Ethnicity Major', 'Progression Rate (%)',
               'vs White British (pp)', 'Position', 'HE Students']
for i, h in enumerate(eth_headers, 1):
    ws1.cell(17, i).value = h
style_header_row(ws1, 17, len(eth_headers))

eth_top = eth.head(15)
for r_idx, (_, row) in enumerate(eth_top.iterrows(), 18):
    ws1.cell(r_idx, 1).value = row['ethnicity_minor']
    ws1.cell(r_idx, 2).value = row['ethnicity_major']
    ws1.cell(r_idx, 3).value = row['progression_rate']
    ws1.cell(r_idx, 4).value = row['gap_vs_white_british_pp']
    ws1.cell(r_idx, 5).value = row['relative_position']
    ws1.cell(r_idx, 6).value = row['number_of_he_students']
    pos = row['relative_position']
    color = GREEN_FILL if pos == 'Above White British' else (RED_FILL if pos == 'Below White British' else AMBER_FILL)
    for c in range(1, 7):
        ws1.cell(r_idx, c).fill = color

auto_width(ws1)

chart1 = BarChart()
chart1.type = 'bar'
chart1.title = 'HE Access Rate by IMD Quintile'
chart1.y_axis.title = 'Access Rate (%)'
chart1.x_axis.title = 'IMD Quintile'
data_ref = Reference(ws1, min_col=2, max_col=2, min_row=9, max_row=14)
cats_ref = Reference(ws1, min_col=1, min_row=10, max_row=14)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.shape = 4
chart1.width = 16
chart1.height = 10
ws1.add_chart(chart1, 'H3')


# ============================================================
# Sheet 2: Attainment Gaps (RAG)
# ============================================================
ws2 = wb.create_sheet('Attainment Gaps (RAG)')

att_eth = pd.read_csv(os.path.join(OUTPUT, 'attainment_gap_ethnicity.csv'))
att_dis = pd.read_csv(os.path.join(OUTPUT, 'attainment_disability.csv'))

ws2['A1'] = 'ATTAINMENT GAPS: RAG-Coded Summary'
ws2['A1'].font = Font(bold=True, size=13)
ws2.merge_cells('A1:G1')

ws2['A3'] = 'Section 1: Good Honours Rate by Ethnicity'
ws2['A3'].font = BOLD

eth_att_hdrs = ['Ethnicity', 'Good Honours (%)', 'White Good Honours (%)',
                'Gap vs White (pp)', 'Gap Category', 'Total Qualifiers', 'Rank']
for i, h in enumerate(eth_att_hdrs, 1):
    ws2.cell(4, i).value = h
style_header_row(ws2, 4, len(eth_att_hdrs))

for r_idx, (_, row) in enumerate(att_eth.iterrows(), 5):
    ws2.cell(r_idx, 1).value = row['ethnicity_code']
    ws2.cell(r_idx, 2).value = row['pct_good_honours']
    ws2.cell(r_idx, 3).value = row['white_pct_good_honours']
    ws2.cell(r_idx, 4).value = row['gap_vs_white_pp']
    ws2.cell(r_idx, 5).value = row['gap_category']
    ws2.cell(r_idx, 6).value = row['total_qualifiers']
    ws2.cell(r_idx, 7).value = row['attainment_rank']
    gap = row['gap_vs_white_pp']
    fill = GREEN_FILL if gap <= 0 else (AMBER_FILL if gap < 15 else RED_FILL)
    for c in range(1, 8):
        ws2.cell(r_idx, c).fill = fill

dis_data_row = 5 + len(att_eth) + 2

ws2.cell(dis_data_row - 1, 1).value = 'Section 2: Good Honours Rate by Disability Type'
ws2.cell(dis_data_row - 1, 1).font = BOLD

dis_hdrs = ['Group', 'Result Type', 'Good Honours (%)', 'Total Qualifiers',
            'Gap vs Non-Disabled (pp)', 'Gap Severity']
for i, h in enumerate(dis_hdrs, 1):
    ws2.cell(dis_data_row, i).value = h
style_header_row(ws2, dis_data_row, len(dis_hdrs))

for r_idx, (_, row) in enumerate(att_dis.iterrows(), dis_data_row + 1):
    ws2.cell(r_idx, 1).value = row['group_label']
    ws2.cell(r_idx, 2).value = row['result_type']
    ws2.cell(r_idx, 3).value = row['good_honours_pct']
    ws2.cell(r_idx, 4).value = row['total_qualifiers'] if pd.notna(row['total_qualifiers']) else ''
    ws2.cell(r_idx, 5).value = row['gap_vs_non_disabled'] if pd.notna(row['gap_vs_non_disabled']) else ''
    ws2.cell(r_idx, 6).value = row['gap_severity'] if pd.notna(row['gap_severity']) else ''
    gap = row['gap_vs_non_disabled']
    if pd.notna(gap):
        fill = GREEN_FILL if abs(gap) < 2 else (AMBER_FILL if abs(gap) < 5 else RED_FILL)
        for c in range(1, 7):
            ws2.cell(r_idx, c).fill = fill

auto_width(ws2)

legend_row = dis_data_row + len(att_dis) + 3
ws2.cell(legend_row, 1).value = 'RAG KEY:'
ws2.cell(legend_row, 1).font = BOLD
ws2.cell(legend_row + 1, 1).value = 'Green'
ws2.cell(legend_row + 1, 1).fill = GREEN_FILL
ws2.cell(legend_row + 1, 2).value = 'No gap / small gap (<5pp)'
ws2.cell(legend_row + 2, 1).value = 'Amber'
ws2.cell(legend_row + 2, 1).fill = AMBER_FILL
ws2.cell(legend_row + 2, 2).value = 'Moderate gap (5-15pp)'
ws2.cell(legend_row + 3, 1).value = 'Red'
ws2.cell(legend_row + 3, 1).fill = RED_FILL
ws2.cell(legend_row + 3, 2).value = 'Large gap (15pp+)'


# ============================================================
# Sheet 3: Graduate Outcomes
# ============================================================
ws3 = wb.create_sheet('Graduate Outcomes')

sal = pd.read_csv(os.path.join(OUTPUT, 'graduate_salary_by_subject.csv'))
dest = pd.read_csv(os.path.join(OUTPUT, 'graduate_destinations.csv'))

ws3['A1'] = 'GRADUATE OUTCOMES: Employment by Subject'
ws3['A1'].font = Font(bold=True, size=13)
ws3.merge_cells('A1:G1')

ws3['A3'] = 'Section 1: Full-Time Employment Rate by Subject Area'
ws3['A3'].font = BOLD

sal_hdrs = ['Rank', 'Subject Area', 'FT Employment (%)', 'Total Graduates',
            'Quartile', 'Outcome Rating']
for i, h in enumerate(sal_hdrs, 1):
    ws3.cell(4, i).value = h
style_header_row(ws3, 4, len(sal_hdrs))

quartile_fills = {
    'Top quartile': GREEN_FILL,
    'Upper-middle quartile': PatternFill('solid', fgColor='D5F0D5'),
    'Lower-middle quartile': AMBER_FILL,
    'Bottom quartile': RED_FILL,
}
for r_idx, (_, row) in enumerate(sal.iterrows(), 5):
    ws3.cell(r_idx, 1).value = row['ft_employment_rank']
    ws3.cell(r_idx, 2).value = row['subject_area_of_degree']
    ws3.cell(r_idx, 3).value = row['pct_ft_employed']
    ws3.cell(r_idx, 4).value = row['total_graduates']
    ws3.cell(r_idx, 5).value = row['employment_quartile']
    ws3.cell(r_idx, 6).value = row['outcome_rating']
    fill = quartile_fills.get(row['employment_quartile'])
    if fill:
        for c in range(1, 7):
            ws3.cell(r_idx, c).fill = fill

dest_row = 5 + len(sal) + 2
ws3.cell(dest_row - 1, 1).value = 'Section 2: Graduate Destinations by Subject'
ws3.cell(dest_row - 1, 1).font = BOLD

dest_hdrs = ['Rank', 'Subject Area', 'Any Employment (%)', 'FT Employment (%)',
             'Further Study (%)', 'Unemployed (%)', 'Total Known', 'Employment Band']
for i, h in enumerate(dest_hdrs, 1):
    ws3.cell(dest_row, i).value = h
style_header_row(ws3, dest_row, len(dest_hdrs))

dest_clean = dest[dest['subject_area_of_degree'] != 'Unknown subject']
for r_idx, (_, row) in enumerate(dest_clean.iterrows(), dest_row + 1):
    ws3.cell(r_idx, 1).value = row['employment_rank']
    ws3.cell(r_idx, 2).value = row['subject_area_of_degree']
    ws3.cell(r_idx, 3).value = row['pct_any_employment']
    ws3.cell(r_idx, 4).value = row['pct_ft_employed']
    ws3.cell(r_idx, 5).value = row['pct_any_study']
    ws3.cell(r_idx, 6).value = row['pct_unemployed']
    ws3.cell(r_idx, 7).value = row['total_known']
    ws3.cell(r_idx, 8).value = row['employment_band']
    band = row['employment_band']
    fill = GREEN_FILL if band == 'High employment' else (AMBER_FILL if band == 'Moderate employment' else RED_FILL)
    for c in range(1, 9):
        ws3.cell(r_idx, c).fill = fill

auto_width(ws3)


# ============================================================
# Sheet 4: Provider Comparison
# ============================================================
ws4 = wb.create_sheet('Provider Comparison')

ws4['A1'] = 'PROVIDER COMPARISON: Access, Attainment and Progression'
ws4['A1'].font = Font(bold=True, size=13)
ws4.merge_cells('A1:H1')

scorecard_path = os.path.join(OUTPUT, 'provider_scorecard.csv')
prov_path = os.path.join(OUTPUT, 'attainment_by_provider.csv')

if os.path.exists(scorecard_path):
    sc = pd.read_csv(scorecard_path)
    ws4['A3'] = 'Provider Scorecard: Multi-dimensional Ranking'
    ws4['A3'].font = BOLD

    sc_hdrs = ['Provider', 'Access Rate (%, IMD Q1)', 'Good Honours (%)',
               'Progression (%)', 'Access Rank', 'Attainment Rank',
               'Progression Rank', 'Overall Rank', 'Band']
    for i, h in enumerate(sc_hdrs, 1):
        ws4.cell(4, i).value = h
    style_header_row(ws4, 4, len(sc_hdrs))

    band_fills = {
        'Top 20%': GREEN_FILL,
        'Top 40%': PatternFill('solid', fgColor='D5F0D5'),
        'Middle': PatternFill('solid', fgColor='FFFACD'),
        'Lower 40%': AMBER_FILL,
        'Bottom 20%': RED_FILL,
    }
    sc_top50 = sc.head(50)
    for r_idx, (_, row) in enumerate(sc_top50.iterrows(), 5):
        ws4.cell(r_idx, 1).value = row['provider_name']
        ws4.cell(r_idx, 2).value = row['access_rate_pct']
        ws4.cell(r_idx, 3).value = row['good_honours_pct']
        ws4.cell(r_idx, 4).value = row['progression_pct']
        ws4.cell(r_idx, 5).value = row['access_rank']
        ws4.cell(r_idx, 6).value = row['attainment_rank']
        ws4.cell(r_idx, 7).value = row['progression_rank']
        ws4.cell(r_idx, 8).value = row['overall_rank']
        ws4.cell(r_idx, 9).value = row['overall_band']
        fill = band_fills.get(row['overall_band'])
        if fill:
            for c in range(1, 10):
                ws4.cell(r_idx, c).fill = fill
else:
    ws4['A3'] = 'Provider scorecard data will be available once ap_providers table is fully loaded.'
    ws4['A3'].font = Font(italic=True)

if os.path.exists(prov_path):
    prov = pd.read_csv(prov_path)
    prov_row = 5 + (len(sc_top50) + 3 if os.path.exists(scorecard_path) else 3)
    ws4.cell(prov_row, 1).value = 'Provider Attainment Distribution (top 20 by good honours rate)'
    ws4.cell(prov_row, 1).font = BOLD

    prov_hdrs = ['Rank', 'Provider', 'Good Honours (%)',
                 'Sector Average (%)', 'Deviation (pp)', 'Performance Band', 'Qualifiers']
    for i, h in enumerate(prov_hdrs, 1):
        ws4.cell(prov_row + 1, i).value = h
    style_header_row(ws4, prov_row + 1, len(prov_hdrs))

    prov_top20 = prov.head(20)
    for r_idx, (_, row) in enumerate(prov_top20.iterrows(), prov_row + 2):
        ws4.cell(r_idx, 1).value = row['rank_highest']
        ws4.cell(r_idx, 2).value = row['provider_name']
        ws4.cell(r_idx, 3).value = row['good_honours_rate']
        ws4.cell(r_idx, 4).value = row['sector_average']
        ws4.cell(r_idx, 5).value = row['deviation_from_avg']
        ws4.cell(r_idx, 6).value = row['performance_band']
        ws4.cell(r_idx, 7).value = row['total_qualifiers']
        dev = row['deviation_from_avg']
        fill = GREEN_FILL if dev >= 5 else (RED_FILL if dev <= -5 else AMBER_FILL)
        for c in range(1, 8):
            ws4.cell(r_idx, c).fill = fill

auto_width(ws4)

wb.save(EXCEL_PATH)
print(f'Saved {EXCEL_PATH}')
